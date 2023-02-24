import openpyxl 
import string
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import cell
# from openpyxl.compat import range
from openpyxl.utils import get_column_letter

import xlsxwriter
import matplotlib.pyplot as plot
from os import listdir, walk
from os.path import isfile, join
import tkinter as tk
from tkinter import *

def copyOBAdata(srs_wb,report_entry):
    srs_OBA_data = srs_wb['OBA']
    # need to write a function here to copy OBA data for each:
    # report_dataentry = report_file['ASTC Source']
    overalloct = list()
    temp_overalloct = srs_OBA_data['B3:M3']
    maxoct = list()
    temp_Maxoct = srs_OBA_data['B4:M4']
    minoct = list()
    temp_minoct = srs_OBA_data['B5:M5']

    thirdoverall = list()
    temp_thirdoverall = srs_OBA_data['B9:AK9']
    maxoveral = list()
    temp_maxoveral = srs_OBA_data['B10:AK10']
    minoveral = list()
    temp_minoveral = srs_OBA_data['B11:AK11']
    for row in temp_overalloct:
        for num in row:
            overalloct.append(num.value)
    for row in temp_Maxoct:
        for num in row:
            maxoct.append(num.value)
    for row in temp_minoct:
        for num in row:
            minoct.append(num.value)

    for row in temp_thirdoverall:
        for num in row:
            thirdoverall.append(num.value)
    for row in temp_maxoveral:
        for num in row:
            maxoveral.append(num.value)
    for row in temp_minoveral:
        for num in row:
            minoveral.append(num.value)
    report_entry.append(overalloct)
    report_entry.append(maxoct)
    report_entry.append(minoct)
    report_entry.append(thirdoverall)
    report_entry.append(maxoveral)
    report_entry.append(minoveral)



STC_contour = [28,31,34,37,40,43,44,45,46,47,48,48,48,48,48,48]

# testplan data import - list of all tests and their associated SLM data file names
# HARDCODED -- CHANGE TO GUI entry
testplan_path ='INCOMING - PAFWC STC Testing Plan List_SRBR.xlsx'
testplanfile = openpyxl.load_workbook(testplan_path)
# copy all entries from  testplan and put them in a list 
testlist = testplanfile.active

# testdata.test_num = testlist['C'] # need to remove 'None' Values..

test_num = testlist['C6:C10']
source_slm_num = testlist['K6:K10']
receive_slm_num = testlist['L6:L10']
bkgnd_slm_num = testlist['M6:M10']
rt_slm_num = testlist['N6:N10']
ASTC_enable = testlist['O']
NIC_enable = testlist['P']


# HARDCODED -- CHANGE TO GUI entry
# path for SLM raw data - use forward slash for windows paths
rawDtestpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Test Data/Raw Data/SLM D/'
rawEtestpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Test Data/Raw Data/SLM E/'


D_datafiles = [f for f in listdir(rawDtestpath) if isfile(join(rawDtestpath,f))]
E_datafiles = [f for f in listdir(rawEtestpath) if isfile(join(rawEtestpath,f))]
# need to differentiate D and E files somehow...maybe append the raw path into a tuple

rawReportpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Documents/Python_devstuff/Example_folder/'

reports =  [f for f in listdir(rawReportpath) if isfile(join(rawReportpath,f))]

# slm_file = openpyxl.load_workbook(rawDtestpath+D_datafiles[1])

Source_room_vol = 10080  # in cubic ft 
Recieve_room_vol = 12096 # in cubic ft
# write a function for getting the ASTC number from the data collected
df = pd.read_excel(rawReportpath+reports[1], sheet_name='SLM Data')
