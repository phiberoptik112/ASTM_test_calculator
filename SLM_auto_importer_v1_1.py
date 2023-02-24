import openpyxl 
import string
# import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import cell
# from openpyxl.compat import range
from openpyxl.utils import get_column_letter

import xlsxwriter
import matplotlib.pyplot as plot
from os import listdir, walk
from os.path import isfile, join
import tkinter as tk
from tkinter import *

# testdata = dict()
# class testdata {
#     'test_num',
#     'source_slm_num',
#     'receive_slm_num',
#     'bkgnd_slm_num',
#     'rt_slm_num',
#     'ASTC',
#     'NIC'}


STC_contour = [28, 31,34,37,40,43,44,45,46,47,48,48,48,48,48,48]
# testcount = 36 # i dont like hardcoding test count. should come from the file

# testplan data import - list of all tests and their associated SLM data file names
# HARDCODED -- CHANGE TO GUI entry
testplan_path ='INCOMING - PAFWC STC Testing Plan List_SRBR.xlsx'
testplanfile = openpyxl.load_workbook(testplan_path)
# copy all entries from  testplan and put them in a list 
testlist = testplanfile.active
# K column, index 5 is start of tests
testnum_index = 5
# testdata.test_num = testlist['C'] # need to remove 'None' Values..
test_num = testlist['C']
source_slm_num = testlist['K']
receive_slm_num = testlist['L']
bkgnd_slm_num = testlist['M']
rt_slm_num = testlist['N']
ASTC_enable = testlist['O']
NIC_enable = testlist['P']


# HARDCODED -- CHANGE TO GUI entry
# path for SLM raw data - use forward slash for windows paths
rawDtestpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Test Data/Raw Data/SLM D/'
rawEtestpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Test Data/Raw Data/SLM E/'


D_datafiles = [f for f in listdir(rawDtestpath) if isfile(join(rawDtestpath,f))]
E_datafiles = [f for f in listdir(rawEtestpath) if isfile(join(rawEtestpath,f))]
# need to differentiate D and E files somehow...maybe append the raw path into a tuple

rawReportpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Documents/Python_devstuff/Example_folder/'

reports =  [f for f in listdir(rawReportpath) if isfile(join(rawReportpath,f))]

slm_file = openpyxl.load_workbook(rawDtestpath+D_datafiles[1])

for i in range(len(reports)):
    


report_file = openpyxl.load_workbook(rawReportpath+reports[1])
# report_file = Workbook()
targetsheet = 'ASTC Source'
targetsheet = report_file.active

OBA_data = slm_file['OBA']

# report_dataentry = report_file['ASTC Source']
report_dataentry = report_file.create_sheet('ASTC Source auto')
report_recieveentry = report_file.create_sheet('ASTC Recieve auto')
report_bkgrndentry = report_file.create_sheet('ASTC Bkgnd auto')
report_rtentry = report_file.create_sheet('ASTC RT auto')

overalloct = list()
temp_overalloct = OBA_data['B3:M3']
maxoct = list()
temp_Maxoct = OBA_data['B4:M4']
minoct = list()
temp_minoct = OBA_data['B5:M5']

thirdoverall = list()
temp_thirdoverall = OBA_data['B9:AK9']
maxoveral = list()
temp_maxoveral = OBA_data['B10:AK10']
minoveral = list()
temp_minoveral = OBA_data['B11:AK11']


for row in temp_overalloct:
    for num in row:
        overalloct.append(num.value)
for row in temp_Maxoct:
    for num in row:
        maxoct.append(num.value)
for row in temp_minoct:
    for num in row:
        minoct.append(num.value)

report_dataentry.append(overalloct)
report_dataentry.append(maxoct)
report_dataentry.append(minoct)

report_file.save(filename = rawReportpath+reports[1])


# temp = report_dataentry['B3']

# for i in range(len(report_datafile)):
#     for num in row:
#     report_datafile[i]


# # finding SLM datafile corresponding to  test number 
# for entry in test_num:
# ##check if None, skip
#     currsrs = source_slm_num[entry].value 
#     ##parse for D or E datafiles
#     datafile_num = currsrs[1:]
#     ##D meter, pull D_datafiles
#     ## (D_datafiles) for filename
#     slm_file = [x for x in all_datafiles if currsrs in x]
#     # for x in D_datafiles:
#     #     if currsrs in x:
#     #         slm_file = x
#     Dslm_file = openpyxl.load_workbook(rawDtestpath+D_datafiles[slm_file])
#     Dslm_file = slm_file.active
#     ##store data from OBS worksheet into temp variables
#     temp_sourcevar = Dslm_file['B:M']

## find receive file (in the same way)
