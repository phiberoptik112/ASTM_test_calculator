import openpyxl 
import string
import pandas as pd
import numpy as np
from openpyxl import load_workbook
from openpyxl import Workbook
from openpyxl import cell
from openpyxl.utils import get_column_letter
import xlsxwriter
import matplotlib.pyplot as plot
from os import listdir, walk
from os.path import isfile, join
import tkinter as tk
from tkinter import *

# define a function to pass all the slm number lists into to assign data to reports
#either that or just loop it first and write it later hah
def copyOBAdata(srs_wb,report_entry):
    srs_OBA_data = srs_wb['OBA'] # worksheet def
    overalloct = list()
    temp_overalloct = srs_OBA_data['B3:M3']
    maxoct = list()
    temp_Maxoct = srs_OBA_data['B4:M4']
    minoct = list()
    temp_minoct = srs_OBA_data['B5:M5']
    thirdoverall = list()
    temp_thirdoverall = srs_OBA_data['B9:AK9']
    maxoveral = list()
    temp_maxoveral = srs_OBA_data['B10:AK10']
    minoveral = list()
    temp_minoveral = srs_OBA_data['B11:AK11']
    for row in temp_overalloct:
        for num in row:
            overalloct.append(num.value)
    for row in temp_Maxoct:
        for num in row:
            maxoct.append(num.value)
    for row in temp_minoct:
        for num in row:
            minoct.append(num.value)
    for row in temp_thirdoverall:
        for num in row:
            thirdoverall.append(num.value)
    for row in temp_maxoveral:
        for num in row:
            maxoveral.append(num.value)
    for row in temp_minoveral:
        for num in row:
            minoveral.append(num.value)

    report_entry.append(overalloct)
    report_entry.append(maxoct)
    report_entry.append(minoct)
    report_entry.append(thirdoverall)
    report_entry.append(maxoveral)
    report_entry.append(minoveral)

def copyRTdata(srs_wb, report_entry):
    srs_RT_data = srs_wb['Summary']
    rt_data = list()
    temp_rt_data = srs_RT_data['K26:K43']
    for row in temp_rt_data:
        for num in row:
            rt_data.append(num.value)
    report_entry.append(rt_data)

# HARDCODED -- CHANGE TO GUI entry
testplan_path ='INCOMING - PAFWC STC Testing Plan List_SRBR.xlsx'
testplanfile = openpyxl.load_workbook(testplan_path)
# copy all entries from  testplan and put them in a list 
testlist = testplanfile.active
test_num = testlist['C6:C10']
source_slm_num = testlist['K6:K10']
receive_slm_num = testlist['L6:L10']
bkgnd_slm_num = testlist['M6:M10']
rt_slm_num = testlist['N6:N10']
ASTC_enable = testlist['O']
NIC_enable = testlist['P']

# make a panda with all these lists in it to iterate through row-wise
# HARDCODED -- CHANGE TO GUI entry
# path for SLM raw data - use forward slash for windows paths
rawDtestpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Test Data/Raw Data/SLM D/'
rawEtestpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Test Data/Raw Data/SLM E/'
D_datafiles = [f for f in listdir(rawDtestpath) if isfile(join(rawDtestpath,f))]
E_datafiles = [f for f in listdir(rawEtestpath) if isfile(join(rawEtestpath,f))]

rawReportpath = '//DLA-04/Shared/KAILUA PROJECTS/2023/23-001 Hickam AFB Acoustical Testing/Documents/Python_devstuff/Example_folder/'
reports =  [f for f in listdir(rawReportpath) if isfile(join(rawReportpath,f))]

# finding SLM datafile corresponding to  test number 
testnumlist = list()
Src_list = list()
recieve_list = list()
bkgnd_list = list()
rt_list = list()

for entry in test_num:
    for num in entry:
        testnumlist.append(num.value)
for entry in source_slm_num:
    for num in entry:
        Src_list.append(num.value)
for entry in receive_slm_num:
    for num in entry:
        recieve_list.append(num.value)
for entry in bkgnd_slm_num:
    for num in entry:
        bkgnd_list.append(num.value)
for entry in rt_slm_num:
    for num in entry:
        rt_list.append(num.value)
# for i in range(len(testlist)):
srs_slm_file = list()
receive_slm_file = list()
bkgrnd_slm_file = list()
rt_slm_file = list()
#finding source files
for i in range(len(testnumlist)):
    currsrs = Src_list[i]
    currtest = testnumlist[i]
        ##parse for D or E datafiles
    if currsrs[0] == 'D':
        datafile_num = currsrs[1:]
        ##D meter, pull D_datafiles
            ## (D_datafiles) for filename
        datafile_num = datafile_num+'.xlsx'
        srs_slm_found = [x for x in D_datafiles if datafile_num in x]
        # print(srs_slm_found)
        srs_slm_file.append(rawDtestpath+srs_slm_found[0])
    elif currsrs[0] == 'E':
        datafile_num = currsrs[1:]
        ##E meter, pull E_datafiles
        datafile_num = datafile_num+'.xlsx'
        srs_slm_found = [x for x in E_datafiles if datafile_num in x]
        # print(rec_slm_found)
        receive_slm_file.append(rawEtestpath+srs_slm_found[0])
#finding recieve files
for i in range(len(testnumlist)):
    currrec = recieve_list[i]
    currtest = testnumlist[i]
    if currrec[0] == 'D':
        datafile_num = currrec[1:]
        ##D meter, pull D_datafiles
            ## (D_datafiles) for filename
        datafile_num = datafile_num+'.xlsx'
        rec_slm_found = [x for x in D_datafiles if datafile_num in x]
        # print(rec_slm_found)
        receive_slm_file.append(rawDtestpath+rec_slm_found[0])
    elif currrec[0] == 'E':
        datafile_num = currrec[1:]
        ##E meter, pull E_datafiles
        datafile_num = datafile_num+'.xlsx'
        rec_slm_found = [x for x in E_datafiles if datafile_num in x]
        # print(rec_slm_found)
        receive_slm_file.append(rawEtestpath+rec_slm_found[0])
# print(srs_slm_file)
# print(receive_slm_file)

#finding background files
for i in range(len(testnumlist)):
    currrec = bkgnd_list[i]
    currtest = testnumlist[i]
    if currrec == 'None':continue
        ##parse for D or E datafiles
    if currrec[0] == 'D':
        datafile_num = currrec[1:]

        ##D meter, pull D_datafiles
            ## (D_datafiles) for filename
        datafile_num = datafile_num+'.xlsx'
        bkgrnd_slm_found = [x for x in D_datafiles if datafile_num in x]
        bkgrnd_slm_file.append(rawDtestpath+bkgrnd_slm_found[0])
    elif currrec[0] == 'E':
        datafile_num = currrec[1:]
        ##E meter, pull E_datafiles
        datafile_num = datafile_num+'.xlsx'
        bkgrnd_slm_found = [x for x in E_datafiles if datafile_num in x]
        bkgrnd_slm_file.append(rawEtestpath+bkgrnd_slm_found[0])

# print(bkgrnd_slm_file)
#finding RT files
for i in range(len(testnumlist)):
    currrec = rt_list[i]
    currtest = testnumlist[i]
    if currrec == 'None':continue
        ##parse for D or E datafiles
    if currrec[0] == 'D':
        datafile_num = currrec[1:]
        # print(datafile_num)
        ##D meter, pull D_datafiles
            ## (D_datafiles) for filename
        datafile_num = datafile_num+'.xlsx'
        rt_slm_found = [x for x in D_datafiles if datafile_num in x]
        # print(rec_slm_found)
        rt_slm_file.append(rawDtestpath+rt_slm_found[0])
    elif currrec[0] == 'E':
        datafile_num = currrec[1:]
        # print(datafile_num)
        ##E meter, pull E_datafiles
        datafile_num = datafile_num+'.xlsx'
        rt_slm_found = [x for x in E_datafiles if datafile_num in x]
        # print(rec_slm_found)
        rt_slm_file.append(rawEtestpath+rt_slm_found[0])

# print(rt_slm_file)
## need to check if multiple entries 

for i in range(len(reports)):
    #must order the test file templates in same order as test plan file,
    #or put in logic to find test number and match it with test list
    report_file = openpyxl.load_workbook(rawReportpath+reports[i])
    curr_srs = openpyxl.load_workbook(srs_slm_file[i])
    curr_recive = openpyxl.load_workbook(receive_slm_file[i])
    curr_bkgrd = openpyxl.load_workbook(bkgrnd_slm_file[i])
    curr_rt = openpyxl.load_workbook(rt_slm_file[i])

    report_srsentry = report_file.create_sheet('ASTC Source auto')
    report_recieveentry = report_file.create_sheet('ASTC Recieve auto')
    report_bkgrndentry = report_file.create_sheet('ASTC Bkgnd auto')
    report_rtentry = report_file.create_sheet('ASTC RT auto')

    srs_OBA_data = curr_srs['OBA']
    recive_OBA_data = curr_recive['OBA']
    bkgrnd_OBA_data = curr_bkgrd['OBA']
    rt_OBA_data = curr_rt['Summary']

    copyOBAdata(curr_srs,report_srsentry)
    copyOBAdata(curr_recive,report_recieveentry)
    copyOBAdata(curr_bkgrd,report_bkgrndentry)
    copyRTdata(curr_rt,report_rtentry)
    report_file.save(filename = rawReportpath+reports[i])
    print('Saving: ',rawReportpath+reports[i])


### #### #### working v1 above - with template, still some work to process ASTC
Source_room_vol = 10080  # in cubic ft 
Recieve_room_vol = 12096 # in cubic ft
# write a function for getting the ASTC number from the data collected
df = pd.read_excel(rawReportpath+reports[1], sheet_name='SLM Data')
STC_contour = [28,31,34,37,40,43,44,45,46,47,48,48,48,48,48,48]
# ATL = 


# stc_deficiencies = ATL - STC_contour
# stc_positive_defic= 